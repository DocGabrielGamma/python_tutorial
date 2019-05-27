#!/usr/bin/python
import sys, getopt
from pathlib import Path
from openpyxl import Workbook, load_workbook
from Modules.excelReader import openExcelAt, extractAllRowsWithGroup
from Modules.excelWriter import createFileByGroups
from Modules import excelReader, excelWriter, pathReader
import Modules.Constants as constants

def appendRowsToTable(inputFile, excelFilesByGroups, offsetRow = 0, offsetColumn = 1):
    for group in excelFilesByGroups:
        rows = excelFilesByGroups[group][constants.ROWS_KEY]
        workbook = excelFilesByGroups[group][constants.WORKBOOK_KEY]
        path = excelFilesByGroups[group][constants.PATH_KEY]
        sheet = workbook.active
        initialRow = sheet.max_row
        for rowIndex, keys in enumerate(rows):
            for columnIndex, value in enumerate(rows[keys]):
                rowCoordinate = rowIndex + initialRow + offsetRow
                columnCoordinate = columnIndex + offsetColumn
                sheet.cell(row=rowCoordinate, column=columnCoordinate, value=value)
        workbook.save(path)
    return "Success"

def createExcelFileForGroup(inputFiles, outputPath, columnHeader):
    for file in inputFiles:
        groups = excelReader.getDifferentGroupsFile(file, columnHeader)
        excelFilesByGroups = createFileByGroups(groups, file, outputPath)
        excelFilesByGroups = extractAllRowsWithGroup(file, excelFilesByGroups)
        appendRowsToTable(file, excelFilesByGroups)
    return groups

def main(arguments):
    inputPath, outputPath, columnHeader = pathReader.getFilePaths(arguments)
    inputExcelFiles = excelReader.openExcelAt(inputPath)
    createExcelFileForGroup(inputExcelFiles, outputPath, columnHeader)
    print(inputPath)
    print(outputPath)

if __name__ == "__main__":
    # Give me all the arguments if the program is the main
    # Arguments from 1 to size
    main(sys.argv[1:])
