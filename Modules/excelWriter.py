from openpyxl import Workbook, load_workbook
from Modules.excelReader import getExcelFileNamesinPath
from os.path import isfile, join
import Modules.Constants as constants

def createExcelInPath(path):
    wb = Workbook()
    wb.save(path)
    return wb

def createFileIfRequired(name, outputPath):
    excelName = name + ".xlsx"
    excelFilesNames = getExcelFileNamesinPath(outputPath)
    finalPath = join(outputPath, excelName)
    if len(excelFilesNames) < 1:
        return createExcelInPath(finalPath), finalPath
    for excelFileName in excelFilesNames:
        if excelFileName.find(excelName) > -1:
           return load_workbook(finalPath), finalPath
    return createExcelInPath(finalPath), finalPath

def createFileByGroups(groups, file, outputPath):
    excelFilesByGroups = {}
    for group in groups:
      excelFile, path = createFileIfRequired(group, outputPath)
      fileProperties = {
          constants.WORKBOOK_KEY: excelFile,
          constants.PATH_KEY: path,
          constants.ROWS_KEY: None
      }
      excelFilesByGroups[group] = fileProperties
    return excelFilesByGroups

def appendRowsToTable(inputFile, excelFilesByGroups, offsetRow = 0, offsetColumn = 1):
    for group in excelFilesByGroups:
        rows = excelFilesByGroups[group][constants.ROWS_KEY]
        workbook = excelFilesByGroups[group][constants.WORKBOOK_KEY]
        path = excelFilesByGroups[group][constants.PATH_KEY]
        sheet = workbook.active
        initialRow = sheet.max_row
        for rowIndex, keys in enumerate(rows):
            for columnIndex, value in enumerate(rows[keys]):
                rowCoordinate = rowIndex + initialRow + offsetRow
                columnCoordinate = columnIndex + offsetColumn
                sheet.cell(row=rowCoordinate, column=columnCoordinate, value=value)
        workbook.save(path)
    return "Success"