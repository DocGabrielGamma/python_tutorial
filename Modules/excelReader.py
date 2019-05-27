from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook
import Modules.Constants as constants

def isExcelFileName(completePath):
    if isfile(completePath) and completePath.endswith('.xlsx'):
        return True
    return False

def constructFileNamesList(excelFiles, file, path):
    completePath = join(path, file)
    if isExcelFileName(completePath):
        excelFiles.append(completePath)
    return excelFiles

def getExcelFileNamesinPath(path):
    excelFileNames = []
    #return another option [files for files in listdir(path) if isfile(join(path, files))]
    files = listdir(path)
    if len(files) > 0: 
        for file in files:
            excelFileNames = constructFileNamesList(excelFileNames, file, path)
        return excelFileNames
    return []

def constructFileList(excelFilenames):
    excelFiles = []
    for filename in excelFilenames:
        excelFiles.append(load_workbook(filename))
    return excelFiles

def openExcelAt(path):
    excelFilenames = getExcelFileNamesinPath(path)
    return constructFileList(excelFilenames)

def findColumnWithValue(columnHeader, sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == columnHeader:
                return cell.column_letter

def findRowWithValue(value, sheet):
    rows = {}
    for row in sheet.iter_rows():
        values = []
        for cell in row:
            values.append(cell.value)
            if str((cell.value)) == value:
                rows[cell.row] = values
    return rows

def addToArrayifUnique(array, value, columnHeader):
    valueToIntroduce = ""
    if type(value) != str:
        valueToIntroduce = str((value))
    else:
        valueToIntroduce = value
    if valueToIntroduce not in array and not valueToIntroduce == columnHeader:
        array.append(valueToIntroduce)
    return array

def getDifferentGroupsFile(file, columnHeader):
    sheet = file.active
    groups = []
    columnLetter = findColumnWithValue(columnHeader , sheet)
    for cell in sheet[columnLetter]:
        groups = addToArrayifUnique(groups, cell.value, columnHeader)
    return groups

def extractAllRowsWithGroup(file, excelFilesByGroups):
    for group in excelFilesByGroups:
        excelFilesByGroups[group][constants.ROWS_KEY] = findRowWithValue(group, file.active)
    return excelFilesByGroups