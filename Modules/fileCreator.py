from openpyxl import Workbook, load_workbook
from Modules.excelReader import getExcelFileNamesinPath

def createExcelInPath(path):
    wb = Workbook()
    wb.save(path)
    return wb

def createFileIfRequired(name, outputPath):
    excelName = name + ".xlsx"
    excelFilesNames = getExcelFileNamesinPath(outputPath)
    finalPath = outputPath + '\\' + excelName
    if len(excelFilesNames) < 1:
        return createExcelInPath(finalPath), finalPath
    for excelFileName in excelFilesNames:
        if excelFileName.find(excelName) > -1:
           return load_workbook(finalPath), finalPath
    return createExcelInPath(finalPath), finalPath